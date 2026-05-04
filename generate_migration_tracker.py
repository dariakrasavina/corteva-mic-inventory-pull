#!/usr/bin/env python3
"""
Corteva MIC — Migration Tracker Generator (workspace inventory)
===============================================================
Reads workspace_inventory_sdk output and generates one Excel workbook per
environment, with one sheet per asset type and migration decision columns.

─── Output ──────────────────────────────────────────────────────────────────
    tracker_<env>.xlsx   — one file per environment, one sheet per asset type

─── Usage ───────────────────────────────────────────────────────────────────
    python generate_migration_tracker.py
    python generate_migration_tracker.py --input-dir /path/to/workspace_inventory_sdk
    python generate_migration_tracker.py --output-dir /path/to/output

─── Requirements ────────────────────────────────────────────────────────────
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import glob
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl is required.  pip install openpyxl", file=sys.stderr)
    sys.exit(1)


DEFAULT_INPUT  = os.path.expanduser(
    "~/corteva-mic-workspace-assets-client-version/workspace_inventory_sdk"
)
DEFAULT_OUTPUT = os.path.expanduser(
    "~/corteva-mic-workspace-assets-client-version"
)

ASSET_TYPES = [
    "jobs",
    "pipelines",
    "notebooks",
    "tables",
    "volumes",
    "functions",
    "registered_models",
    "serving_endpoints",
    "apps",
    "experiments",
    "dashboards",
    "repos",
    "genie_spaces",
]

DECISION_COLUMNS = [
    "migration_decision",   # stay / move / both / deprecate
    "notes",
    "owner",
    "priority",             # high / medium / low
]

# Colours
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")  # dark blue
DECISION_FILL = PatternFill("solid", fgColor="E2EFDA")  # light green
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
DECISION_FONT = Font(bold=True, color="375623", size=10)
ALT_ROW_FILL  = PatternFill("solid", fgColor="F2F2F2")  # light grey

COL_WIDTHS = {
    "name":               38,
    "full_name":          48,
    "path":               48,
    "notebook_path":      48,
    "migration_decision": 20,
    "notes":              40,
    "owner":              24,
    "priority":           12,
    "workspace":          24,
}


def _read_csv(path: str) -> list[dict]:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(path, newline="", encoding=encoding) as f:
                return list(csv.DictReader(f))
        except (UnicodeDecodeError, UnicodeError):
            continue
    return []


def _load_env_asset(env_dir: str, asset_type: str) -> list[dict]:
    matches = glob.glob(os.path.join(env_dir, f"*_{asset_type}.csv"))
    rows: list[dict] = []
    for path in matches:
        rows.extend(_read_csv(path))
    return rows


def _write_sheet(wb: openpyxl.Workbook, sheet_name: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit

    if not rows:
        ws.append(["No data found for this asset type"])
        return

    asset_cols = list(rows[0].keys())
    all_cols   = asset_cols + DECISION_COLUMNS

    # Header row
    ws.append(all_cols)
    for col_idx, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col_name in DECISION_COLUMNS:
            cell.fill = DECISION_FILL
            cell.font = DECISION_FONT
        else:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        data = [row.get(c, "") for c in asset_cols] + [""] * len(DECISION_COLUMNS)
        ws.append(data)
        if row_idx % 2 == 0:
            for col_idx in range(1, len(all_cols) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ALT_ROW_FILL

    # Dropdowns
    last_row = len(rows) + 1
    decision_col = get_column_letter(all_cols.index("migration_decision") + 1)
    priority_col = get_column_letter(all_cols.index("priority") + 1)

    dv_decision = DataValidation(type="list", formula1='"stay,move,both,deprecate"',
                                 allow_blank=True, showDropDown=False)
    dv_decision.sqref = f"{decision_col}2:{decision_col}{last_row}"
    ws.add_data_validation(dv_decision)

    dv_priority = DataValidation(type="list", formula1='"high,medium,low"',
                                 allow_blank=True, showDropDown=False)
    dv_priority.sqref = f"{priority_col}2:{priority_col}{last_row}"
    ws.add_data_validation(dv_priority)

    # Column widths
    for col_idx, col_name in enumerate(all_cols, start=1):
        width = COL_WIDTHS.get(col_name, 18)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}1"


def generate_for_env(env: str, env_dir: str, output_dir: str) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    total = 0
    for asset_type in ASSET_TYPES:
        rows = _load_env_asset(env_dir, asset_type)
        _write_sheet(wb, asset_type, rows)
        total += len(rows)
        print(f"    {asset_type:<25} {len(rows):>5} rows", file=sys.stderr)

    out_path = os.path.join(output_dir, f"tracker_{env}.xlsx")
    wb.save(out_path)
    print(f"  -> saved: {out_path}  ({total} total rows)\n", file=sys.stderr)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate migration tracker Excel files from workspace_inventory_sdk output",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--input-dir",  default=DEFAULT_INPUT,
                        help=f"Path to workspace_inventory_sdk folder (default: {DEFAULT_INPUT})")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT,
                        help=f"Output folder for tracker .xlsx files (default: {DEFAULT_OUTPUT})")
    args = parser.parse_args()

    if not os.path.isdir(args.input_dir):
        print(f"ERROR: input directory not found: {args.input_dir}", file=sys.stderr)
        sys.exit(1)

    os.makedirs(args.output_dir, exist_ok=True)

    envs = sorted(d for d in os.listdir(args.input_dir)
                  if os.path.isdir(os.path.join(args.input_dir, d)))

    if not envs:
        print("ERROR: no environment subfolders found in input directory", file=sys.stderr)
        sys.exit(1)

    print(f"\n  Input  : {args.input_dir}", file=sys.stderr)
    print(f"  Output : {args.output_dir}", file=sys.stderr)
    print(f"  Envs   : {', '.join(envs)}\n", file=sys.stderr)

    for env in envs:
        env_dir = os.path.join(args.input_dir, env)
        print(f"  [{env}]", file=sys.stderr)
        generate_for_env(env, env_dir, args.output_dir)

    print("  Done.", file=sys.stderr)


if __name__ == "__main__":
    main()
