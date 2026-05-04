#!/usr/bin/env python3
"""
Corteva MIC — Migration Decision Auto-Populator
================================================
Reads the client's manually-pulled inventory Excel files and UC system table
lineage to pre-fill migration_decision + notes in the tracker workbooks.

Sources:
  clients_manually_pulled_inventory/
    UNITY_CATALOG_TABLE_ASSETS.xlsx  → tables sheet (and schema-level inference)
    DABS_TRACKER.xlsx                → pipelines + jobs sheets
  UC_system_tables_scan/<env>/
    *_system_table_lineage.csv       → notebooks + supplemental job/pipeline inference

Coverage mapping (from client labels):
  SEED  → move        (new tenant)
  CROP  → stay        (current tenant)
  CP    → stay        (Crop Protection, synonym for CROP)
  BOTH  → both        (keep in current + migrate to new)
  TC    → stay        (internal grouping — stays)
  SPIN_CO cleared=YES → move
  SPIN_CO migrate=YES → move
  DELETE_ALL schemas  → deprecate (all tables in that schema)
  DAB Active=NO       → deprecate

Decision lookup order per sheet type:
  tables    : (1) catalog+schema+table exact  (2) schema+table  (3) DELETE_ALL schema
  pipelines : (1) lineage target schemas  (2) pipeline name = schema name
  jobs      : (1) lineage target schemas  (2) DAB exact name match
              (3) DAB fuzzy match (underscore→space word containment)
  notebooks : (1) lineage target schemas
  Already-set cells are never overwritten.

─── Usage ───────────────────────────────────────────────────────────────────
    python populate_migration_decisions.py
    python populate_migration_decisions.py --client-dir /path/to/client_folder
    python populate_migration_decisions.py --tracker-dir /path/to/tracker_folder
    python populate_migration_decisions.py --uc-dir /path/to/UC_system_tables_scan
"""

from __future__ import annotations

import argparse
import csv
import glob
import os
import sys
from collections import defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required.  pip install openpyxl", file=sys.stderr)
    sys.exit(1)


DEFAULT_CLIENT_DIR = os.path.expanduser(
    "~/corteva-mic-workspace-assets-client-version/"
    "clients_manually_pulled_inventory/"
    "Environment Resources_UC Table Migration_DABS TRACKER Info EXCEL"
)
DEFAULT_TRACKER_DIR = os.path.expanduser(
    "~/corteva-mic-workspace-assets-client-version/workspace_inventory_sdk"
)
DEFAULT_UC_DIR = os.path.expanduser(
    "~/corteva-mic-workspace-assets-client-version/UC_system_tables_scan"
)

COVERAGE_MAP: dict[str, str] = {
    "seed": "move",
    "crop": "stay",
    "cp":   "stay",
    "both": "both",
    "tc":   "stay",
}


# ─── helpers ─────────────────────────────────────────────────────────────────

def _cov(value) -> str | None:
    if not value:
        return None
    return COVERAGE_MAP.get(str(value).strip().lower())


def _norm(value) -> str:
    return str(value).strip().lower() if value else ""


def _header_map(ws, header_row=1) -> dict[str, int]:
    """Return {column_name: 1-based column index}, first occurrence wins."""
    result: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value and str(cell.value).strip() not in result:
            result[str(cell.value).strip()] = cell.column
    return result


def _consensus(decisions: set) -> str | None:
    """Collapse a set of decisions to one value."""
    decs = {d for d in decisions if d}
    if not decs:
        return None
    if decs == {"deprecate"}:
        return "deprecate"
    mig = decs - {"deprecate"}
    if not mig:
        return "deprecate"
    if len(mig) == 1:
        return next(iter(mig))
    return "both"   # mixed stay+move → both


# ─── load client decisions ────────────────────────────────────────────────────

def build_table_decisions(
    client_dir: str,
) -> tuple[dict, dict, set[str]]:
    """
    Returns:
        cat_decisions : {(catalog_lc, schema_lc, table_lc): {"decision", "notes"}}
        decisions     : {(schema_lc, table_lc): {"decision", "notes"}}
        deprecate_schemas : {schema_lc}
    """
    path = os.path.join(client_dir, "UNITY_CATALOG_TABLE_ASSETS.xlsx")
    wb = load_workbook(path, read_only=True, data_only=True)

    cat_decisions: dict[tuple, dict] = {}
    decisions: dict[tuple, dict]     = {}
    deprecate_schemas: set[str]       = set()

    def _put_with_cat(catalog, schema, table, decision, notes=None):
        if not schema or not table:
            return
        k = (_norm(catalog), _norm(schema), _norm(table))
        if k not in cat_decisions:
            cat_decisions[k] = {"decision": decision, "notes": str(notes) if notes else ""}

    def _put(schema, table, decision, notes=None):
        if not schema or not table:
            return
        k = (_norm(schema), _norm(table))
        if k not in decisions:
            decisions[k] = {"decision": decision, "notes": str(notes) if notes else ""}

    ws = wb["NEW_CORTEVA_SOURCE"]
    h = _header_map(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        _put_with_cat(
            row[h["Catalog"]     - 1] if "Catalog"     in h else None,
            row[h["schema_name"] - 1],
            row[h["table_name"]  - 1],
            _cov(row[h["Coverage"] - 1]),
            row[h["Notes"] - 1] if "Notes" in h else None,
        )

    ws = wb["NEW_CORTEVA_GOLD"]
    h = _header_map(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        _put(
            row[h["schema_name"] - 1],
            row[h["table_name"]  - 1],
            _cov(row[h["Coverage"] - 1]),
            row[h["Notes"] - 1] if "Notes" in h else None,
        )

    ws = wb["NEW_CORTEVA_STAR_SCHEMAS"]
    h = _header_map(ws)
    notes_col = h.get("NOTES") or h.get("Notes")
    for row in ws.iter_rows(min_row=2, values_only=True):
        _put(
            row[h["Schema"] - 1],
            row[h["Table"]  - 1],
            _cov(row[h["Coverage"] - 1]),
            row[notes_col - 1] if notes_col else None,
        )

    ws = wb["SPIN_CO_SOURCE"]
    h = _header_map(ws)
    cleared_col = h.get("Cleared or Not Cleared")
    for row in ws.iter_rows(min_row=2, values_only=True):
        schema  = row[h["Schema"] - 1]
        table   = row[h["Table"]  - 1]
        cleared = row[cleared_col - 1] if cleared_col else None
        if _norm(cleared) == "yes":
            _put(schema, table, "move")

    ws = wb["SPIN_CO_GOLD"]
    h = _header_map(ws)
    migrate_col = h.get("Migrate")
    for row in ws.iter_rows(min_row=2, values_only=True):
        schema  = row[h["Schema"]  - 1]
        table   = row[h["Table"]   - 1]
        migrate = row[migrate_col  - 1] if migrate_col else None
        cov     = row[h["Coverage"]- 1] if "Coverage" in h else None
        if _norm(migrate) == "yes":
            _put(schema, table, "move")
        elif cov:
            _put(schema, table, _cov(cov))

    ws = wb["DELETE_ALL_AND_DON'T_MIGRATE"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        schema = row[0]
        if schema:
            deprecate_schemas.add(_norm(schema))

    ws = wb["FARMER_MASTER"]
    h = _header_map(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        _put_with_cat(
            row[h["Catalog"]     - 1] if "Catalog"     in h else None,
            row[h["schema_name"] - 1],
            row[h["table_name"]  - 1],
            _cov(row[h["Coverage"] - 1]),
        )

    wb.close()
    return cat_decisions, decisions, deprecate_schemas


def build_schema_decisions(
    cat_decisions: dict,
    decisions: dict,
    deprecate_schemas: set[str],
) -> dict[str, str]:
    """
    Aggregate per-table decisions to schema level.
    Used for pipeline-name matching and lineage-based inference.
    """
    schema_decs: dict[str, set] = defaultdict(set)
    for (schema, _table), v in decisions.items():
        if v["decision"]:
            schema_decs[schema].add(v["decision"])
    for (_cat, schema, _table), v in cat_decisions.items():
        if v["decision"]:
            schema_decs[schema].add(v["decision"])
    for schema in deprecate_schemas:
        schema_decs[schema].add("deprecate")
    return {s: _consensus(d) for s, d in schema_decs.items() if _consensus(d)}


def build_dab_decisions(client_dir: str) -> dict[str, dict]:
    """
    Returns {bundle_key: {"decision", "target_schema", "notes"}}
    Active=NO → decision="deprecate"; Active=YES → decision=None (use target_schema instead).
    """
    path = os.path.join(client_dir, "DABS_TRACKER.xlsx")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["DABs Inventory"]

    dab: dict[str, dict] = {}
    h = _header_map(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        bundle = row[h["Bundle Name"]   - 1] if "Bundle Name"   in h else row[0]
        active = row[h["Active"]        - 1] if "Active"        in h else row[7]
        schema = row[h["Target Schema"] - 1] if "Target Schema" in h else None
        notes  = row[h["Notes"]         - 1] if "Notes"         in h else row[8]
        if not bundle:
            continue
        key = _norm(bundle)
        if key.endswith(".yml"):
            key = key[:-4]
        if key not in dab:
            dab[key] = {
                "decision":      "deprecate" if _norm(active) == "no" else None,
                "target_schema": _norm(schema) if schema else "",
                "notes":         str(notes) if notes else "",
            }

    wb.close()
    return dab


def load_lineage(uc_dir: str, env: str) -> dict[tuple, set[str]]:
    """
    Loads <env>/<PREFIX>_system_table_lineage.csv.
    Returns {(entity_type, entity_id): {target_schema_lc, ...}}
    """
    env_dir = os.path.join(uc_dir, env)
    matches = glob.glob(os.path.join(env_dir, "*_system_table_lineage.csv"))
    entity_schemas: dict[tuple, set] = defaultdict(set)
    for fpath in matches:
        try:
            with open(fpath, encoding="utf-8-sig", newline="") as f:
                for row in csv.DictReader(f):
                    etype  = row.get("entity_type", "")
                    eid    = row.get("entity_id", "")
                    schema = row.get("target_table_schema", "")
                    if etype and eid and schema:
                        entity_schemas[(etype, eid)].add(schema.strip().lower())
        except (FileNotFoundError, IOError):
            pass
    return entity_schemas


# ─── apply decisions ─────────────────────────────────────────────────────────

def _infer_from_schemas(
    schemas: set[str],
    schema_decision_map: dict,
) -> str | None:
    return _consensus({schema_decision_map.get(s) for s in schemas})


def _dab_lookup(
    name: str,
    dab_decisions: dict,
    schema_decision_map: dict,
) -> tuple[str | None, str]:
    """
    Try to find a DAB decision for a job/pipeline name.
    Returns (decision, notes).
    Checks: exact key match, then fuzzy word-containment match.
    """
    name_n = _norm(name)

    # Exact key match
    hit = dab_decisions.get(name_n)
    if hit:
        dec = hit["decision"] or schema_decision_map.get(hit["target_schema"])
        return dec, hit["notes"]

    # Fuzzy: bundle key (underscores→spaces) is contained in job name (underscores→spaces)
    name_words = name_n.replace("_", " ")
    for bundle_key, hit in dab_decisions.items():
        if len(bundle_key) < 8:        # too short — too many false positives
            continue
        bundle_words = bundle_key.replace("_", " ")
        if bundle_words in name_words:
            dec = hit["decision"] or schema_decision_map.get(hit["target_schema"])
            if dec:
                return dec, hit["notes"]

    return None, ""


def _update_sheet(
    ws,
    cat_decisions:      dict,
    decisions:          dict,
    deprecate_schemas:  set,
    dab_decisions:      dict,
    schema_decision_map: dict,
    entity_schemas:     dict,
) -> int:
    h = _header_map(ws)
    dec_col   = h.get("migration_decision")
    notes_col = h.get("notes")
    if not dec_col:
        return 0

    is_table    = "schema_name" in h and "table_name"   in h
    is_job      = "job_id"      in h and "name"         in h
    is_pipeline = "pipeline_id" in h and "name"         in h
    is_notebook = "object_id"   in h

    updated = 0
    for row in ws.iter_rows(min_row=2):
        dec_cell = row[dec_col - 1]
        if dec_cell.value:
            continue

        decision = notes_txt = None

        # ── tables ──────────────────────────────────────────────────────────
        if is_table:
            catalog  = row[h["catalog_name"] - 1].value if "catalog_name" in h else None
            schema   = row[h["schema_name"]  - 1].value
            table    = row[h["table_name"]   - 1].value
            schema_lc = _norm(schema)
            table_lc  = _norm(table)

            if schema_lc in deprecate_schemas:
                decision = "deprecate"
            else:
                if catalog:
                    hit = cat_decisions.get((_norm(catalog), schema_lc, table_lc))
                    if hit:
                        decision, notes_txt = hit["decision"], hit["notes"]
                if not decision:
                    hit = decisions.get((schema_lc, table_lc))
                    if hit:
                        decision, notes_txt = hit["decision"], hit["notes"]

        # ── jobs ─────────────────────────────────────────────────────────────
        elif is_job:
            job_id = str(row[h["job_id"] - 1].value or "")
            name   = row[h["name"]   - 1].value or ""

            # 1. Lineage: job_id → target schemas
            schemas = entity_schemas.get(("JOB", job_id), set())
            decision = _infer_from_schemas(schemas, schema_decision_map)

            # 2. Job name IS a known schema
            if not decision:
                decision = schema_decision_map.get(_norm(name))

            # 3. DAB match (exact + fuzzy)
            if not decision:
                decision, notes_txt = _dab_lookup(name, dab_decisions, schema_decision_map)

        # ── pipelines ────────────────────────────────────────────────────────
        elif is_pipeline:
            pipe_id = str(row[h["pipeline_id"] - 1].value or "")
            name    = row[h["name"]       - 1].value or ""

            # 1. Lineage: pipeline_id → target schemas
            schemas = entity_schemas.get(("PIPELINE", pipe_id), set())
            decision = _infer_from_schemas(schemas, schema_decision_map)

            # 2. Pipeline name IS a known schema (very common for DLT pipelines)
            if not decision:
                decision = schema_decision_map.get(_norm(name))

        # ── notebooks ────────────────────────────────────────────────────────
        elif is_notebook:
            obj_id  = str(row[h["object_id"] - 1].value or "")
            schemas = entity_schemas.get(("NOTEBOOK", obj_id), set())
            decision = _infer_from_schemas(schemas, schema_decision_map)

        if decision:
            dec_cell.value = decision
            if notes_col and notes_txt:
                row[notes_col - 1].value = notes_txt
            updated += 1

    return updated


def populate_tracker(
    tracker_path:        str,
    cat_decisions:       dict,
    decisions:           dict,
    deprecate_schemas:   set,
    dab_decisions:       dict,
    schema_decision_map: dict,
    entity_schemas:      dict,
) -> int:
    print(f"  {os.path.basename(tracker_path)}", file=sys.stderr)
    wb = load_workbook(tracker_path)
    total = 0

    for sheet_name in ("tables", "pipelines", "jobs", "notebooks"):
        if sheet_name not in wb.sheetnames:
            continue
        n = _update_sheet(
            wb[sheet_name],
            cat_decisions, decisions, deprecate_schemas,
            dab_decisions, schema_decision_map, entity_schemas,
        )
        print(f"    {sheet_name:<12} {n:>5} cells updated", file=sys.stderr)
        total += n

    wb.save(tracker_path)
    print(f"  -> saved  ({total} total updates)\n", file=sys.stderr)
    return total


# ─── main ────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Auto-populate migration_decision in tracker Excel files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--client-dir",  default=DEFAULT_CLIENT_DIR)
    parser.add_argument("--tracker-dir", default=DEFAULT_TRACKER_DIR)
    parser.add_argument(
        "--uc-dir", default=DEFAULT_UC_DIR,
        help=f"UC_system_tables_scan folder for lineage CSVs (default: {DEFAULT_UC_DIR})",
    )
    args = parser.parse_args()

    if not os.path.isdir(args.client_dir):
        print(f"ERROR: client-dir not found:\n  {args.client_dir}", file=sys.stderr)
        sys.exit(1)

    print("\n  Loading client inventory ...", file=sys.stderr)
    cat_decisions, decisions, deprecate_schemas = build_table_decisions(args.client_dir)
    dab_decisions = build_dab_decisions(args.client_dir)
    schema_decision_map = build_schema_decisions(cat_decisions, decisions, deprecate_schemas)
    print(f"    catalog-specific decisions  : {len(cat_decisions)}", file=sys.stderr)
    print(f"    broad schema+table decisions: {len(decisions)}", file=sys.stderr)
    print(f"    deprecate schemas           : {len(deprecate_schemas)}", file=sys.stderr)
    print(f"    DAB decisions               : {len(dab_decisions)}", file=sys.stderr)
    print(f"    schema-level decisions      : {len(schema_decision_map)}", file=sys.stderr)

    trackers = sorted(glob.glob(os.path.join(args.tracker_dir, "tracker_*.xlsx")))
    if not trackers:
        print(f"ERROR: no tracker_*.xlsx found in {args.tracker_dir}", file=sys.stderr)
        sys.exit(1)

    print(f"\n  Updating {len(trackers)} tracker(s) ...\n", file=sys.stderr)
    for tracker_path in trackers:
        # Derive env name from filename: tracker_sales-mi-dbw-01-prod.xlsx → sales-mi-dbw-01-prod
        env = os.path.basename(tracker_path).removeprefix("tracker_").removesuffix(".xlsx")
        entity_schemas = load_lineage(args.uc_dir, env)
        print(f"    lineage entities loaded: {len(entity_schemas)}  (env={env})", file=sys.stderr)
        populate_tracker(
            tracker_path,
            cat_decisions, decisions, deprecate_schemas,
            dab_decisions, schema_decision_map, entity_schemas,
        )

    print("  Done.", file=sys.stderr)


if __name__ == "__main__":
    main()
